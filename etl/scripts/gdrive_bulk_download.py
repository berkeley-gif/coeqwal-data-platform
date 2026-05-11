#!/usr/bin/env python3
"""
Bulk-download model run ZIPs and trend report CSVs from Google Shared Drive
using rclone, validate DSS contents, stage to S3, and promote to trigger
extraction.

Prerequisites:
  - rclone installed and configured with a "gdrive" remote pointing to the
    COEQWAL Shared Drive (see README for setup)
  - AWS credentials configured (for S3 access)

Usage:
  # Phase 1: download + validate + stage (safe, no extraction triggered)
  python gdrive_bulk_download.py download \
    --listing ../../reference/COEQWAL_Completed_Scenario_Listing.xlsx \
    --s3-bucket coeqwal-model-run

  # Phase 2a: smoke-test one scenario
  python gdrive_bulk_download.py promote \
    --s3-bucket coeqwal-model-run --scenarios s0020

  # Phase 2b: promote all
  python gdrive_bulk_download.py promote --s3-bucket coeqwal-model-run
"""

import argparse
import csv
import io
import json
import logging
import os
import re
import subprocess
import tempfile
import threading
import zipfile
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import boto3
import openpyxl

# Default output directory for audit CSVs. Gitignored via etl/**/output/.
# Override per-invocation with --output-dir.
DEFAULT_OUTPUT_DIR = Path(__file__).parent / "output"

# ---------------------------------------------------------------------------
# classify_dss.py logic (inlined to avoid fragile cross-directory imports)
# ---------------------------------------------------------------------------
EXCLUDED_SUBFOLDERS = ("archive", "discard", "old", "backup")
GW_BASENAMES = ("cvgroundwaterbudget.dss", "cvgroundwaterout.dss")
SV_TIER3 = "_sv"
SV_TIER2: Tuple[str, ...] = ("statevar", "input")
CAL_TIER3 = "_dv"
CAL_TIER2: Tuple[str, ...] = ("out", "output", "results")

SCENARIO_OVERRIDES = {
    "s0023": {"sv": "coeqwal_s9999_sv_v0.1.2.dss"},
    "s0024": {"sv": "coeqwal_s9999_sv_v0.1.2.dss"},
    "s0030": {"dv": "s0030_dcradjhist_2020lu_noflowreqt_dv_20260126v02.dss"},
    "s0031": {"sv": "coeqwal_s9999_sv_v0.1.14.dss"},
    "s0033": {"sv": "coeqwal_s9999_sv_v0.11.15.dss"},
    "s0037": {"dv": "s0037_dcradjbl_2020lu_priorityfullcwn_dv_v1_20260216.dss"},
    "s0039": {"sv": "coeqwal_s9999_sv_v0.1.4.dss"},
    "s0040": {
        "sv": "coeqwal_s9999_sv_v0.1.4.dss",
        "dv": "s0040_usbralt3_2020lu_deltaout35_dv_v0.2_20251211.dss",
    },
    "s0041": {
        "sv": "coeqwal_s9999_sv_v0.1.4.dss",
        "dv": "s0041_usbralt3_2020lu_deltaout45_dv_v0.2_20251211.dss",
    },
    "s0042": {"sv": "coeqwal_s9999_sv_v0.1.4.dss"},
}


def _norm_for_match(path: str) -> str:
    norm = path.replace("\\", "/").lstrip("./").lower()
    return f"/{norm}/"


def _in_excluded_subfolder(path: str) -> bool:
    parts = path.replace("\\", "/").lower().split("/")
    return any(part in EXCLUDED_SUBFOLDERS for part in parts)


def _pick_simple(
    candidates: List[str], tier3_token: str, tier2_tokens: Tuple[str, ...]
) -> Optional[str]:
    if not candidates:
        return None
    for p in candidates:
        if tier3_token in os.path.basename(p).lower():
            return p
    for p in candidates:
        b = os.path.basename(p).lower()
        if any(tok in b for tok in tier2_tokens):
            return p
    return None


def _pick_by_override(candidates: List[str], required_basename: str) -> Optional[str]:
    for p in candidates:
        if os.path.basename(p).lower() == required_basename:
            return p
    return None


def _selection_reason(
    selected: Optional[str], candidates: List[str],
    tier3_token: str, tier2_tokens: Tuple[str, ...],
    overrides: Dict[str, str], override_key: str,
) -> str:
    """Explain why a particular DSS file was selected."""
    if not selected:
        return "none_found"
    if not candidates:
        return "no_candidates"
    b = os.path.basename(selected).lower()
    if override_key in overrides and b == overrides[override_key]:
        return f"override ({override_key}={overrides[override_key]})"
    if tier3_token in b:
        return f"tier3 ('{tier3_token}' in filename)"
    for tok in tier2_tokens:
        if tok in b:
            return f"tier2 ('{tok}' in filename)"
    return "first_candidate"


def classify_dss_in_zip(dss_paths: List[str], scenario_id: str) -> Dict[str, Any]:
    """Classify DSS files found in a ZIP into SV and DV candidates."""
    sv_candidates: List[str] = []
    cal_candidates: List[str] = []
    skipped: List[str] = []
    classification_method = "folder_structure"

    for p in dss_paths:
        if _in_excluded_subfolder(p):
            skipped.append(p)
            continue
        slug = _norm_for_match(p)
        b = os.path.basename(p).lower()
        if "/dss/input/" in slug:
            sv_candidates.append(p)
        elif "/dss/output/" in slug:
            if b not in GW_BASENAMES:
                cal_candidates.append(p)

    if not sv_candidates and not cal_candidates:
        classification_method = "filename_heuristic"
        for p in dss_paths:
            if _in_excluded_subfolder(p):
                continue
            b = os.path.basename(p).lower()
            if b not in GW_BASENAMES:
                if CAL_TIER3 in b or any(tok in b for tok in CAL_TIER2):
                    cal_candidates.append(p)
            if SV_TIER3 in b or any(tok in b for tok in SV_TIER2):
                sv_candidates.append(p)

    overrides = SCENARIO_OVERRIDES.get(scenario_id, {})

    if "sv" in overrides:
        sv_selected = _pick_by_override(sv_candidates, overrides["sv"])
        if not sv_selected:
            sv_selected = _pick_simple(sv_candidates, SV_TIER3, SV_TIER2)
    else:
        sv_selected = _pick_simple(sv_candidates, SV_TIER3, SV_TIER2)

    if "dv" in overrides:
        dv_selected = _pick_by_override(cal_candidates, overrides["dv"])
        if not dv_selected:
            dv_selected = _pick_simple(cal_candidates, CAL_TIER3, CAL_TIER2)
    else:
        dv_selected = _pick_simple(cal_candidates, CAL_TIER3, CAL_TIER2)

    sv_reason = _selection_reason(
        sv_selected, sv_candidates, SV_TIER3, SV_TIER2, overrides, "sv")
    dv_reason = _selection_reason(
        dv_selected, cal_candidates, CAL_TIER3, CAL_TIER2, overrides, "dv")

    return {
        "sv_candidates": sv_candidates,
        "dv_candidates": cal_candidates,
        "sv_selected": sv_selected,
        "dv_selected": dv_selected,
        "sv_reason": sv_reason,
        "dv_reason": dv_reason,
        "skipped": skipped,
        "classification_method": classification_method,
    }


# ---------------------------------------------------------------------------
# rclone helpers
# ---------------------------------------------------------------------------
log = logging.getLogger("gdrive_bulk_download")

RCLONE_REMOTE = "gdrive"


def rclone_lsjson(folder_id: str, subpath: str = "",
                  dirs_only: bool = False,
                  rclone_remote: Optional[str] = None) -> List[Dict]:
    """List contents of a Drive folder via rclone lsjson.

    If folder_id is a Google Drive folder ID, navigates by ID.
    If folder_id is empty/None, navigates by path (subpath is the full
    Shared Drive path, e.g. "s0011_adjBL_wTUCP/Model_Files/").
    """
    remote = rclone_remote or RCLONE_REMOTE
    target = f"{remote}:{subpath}"
    cmd = ["rclone", "lsjson", target]
    if folder_id:
        cmd.append(f"--drive-root-folder-id={folder_id}")
    if dirs_only:
        cmd.append("--dirs-only")
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    if result.returncode != 0:
        log.warning("rclone lsjson failed: %s", result.stderr.strip())
        return []
    try:
        return json.loads(result.stdout)
    except json.JSONDecodeError:
        log.warning("rclone lsjson returned invalid JSON: %s",
                    result.stdout[:200])
        return []


def rclone_copy_file(folder_id: str, remote_path: str,
                     local_dest_dir: str) -> bool:
    """Download a single file from Drive to a local directory."""
    target = f"{RCLONE_REMOTE}:{remote_path}"
    cmd = [
        "rclone", "copy", target, local_dest_dir,
        f"--drive-root-folder-id={folder_id}",
        "--progress",
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=3600)
    if result.returncode != 0:
        log.error("rclone copy failed: %s", result.stderr.strip())
        return False
    return True


def rclone_cat(folder_id: str, remote_path: str) -> Optional[bytes]:
    """Read a small file from Drive into memory."""
    target = f"{RCLONE_REMOTE}:{remote_path}"
    cmd = [
        "rclone", "cat", target,
        f"--drive-root-folder-id={folder_id}",
    ]
    result = subprocess.run(cmd, capture_output=True, timeout=300)
    if result.returncode != 0:
        log.error("rclone cat failed: %s", result.stderr.decode().strip())
        return None
    return result.stdout


# ---------------------------------------------------------------------------
# Excel reader
# ---------------------------------------------------------------------------
RE_FOLDER_ID = re.compile(r"/folders/([A-Za-z0-9_-]+)")


def read_scenario_listing(
    listing_path: str, sheet_name: str = "HistHydro_20260223"
) -> List[Dict[str, str]]:
    """Read scenario listing from the Excel file, extracting hyperlinks."""
    wb = openpyxl.load_workbook(listing_path)
    ws = wb[sheet_name]
    scenarios = []
    for row in range(3, ws.max_row + 1):
        short_code = ws.cell(row=row, column=3).value
        if not short_code or not str(short_code).strip():
            continue
        short_code = str(short_code).strip()

        study_cell = ws.cell(row=row, column=4)
        study_name = str(study_cell.value or "").strip()
        hyperlink = study_cell.hyperlink
        drive_url = hyperlink.target if hyperlink else ""

        folder_id = ""
        if drive_url:
            m = RE_FOLDER_ID.search(drive_url)
            if m:
                folder_id = m.group(1)

        run_date = str(ws.cell(row=row, column=6).value or "").strip()
        notes = str(ws.cell(row=row, column=7).value or "").strip()

        scenarios.append({
            "short_code": short_code,
            "study_name": study_name,
            "drive_url": drive_url,
            "drive_folder_id": folder_id,
            "run_date": run_date,
            "notes": notes,
        })
    wb.close()
    return scenarios


# ---------------------------------------------------------------------------
# ZIP validation
# ---------------------------------------------------------------------------
def validate_zip(zip_path: str, scenario_id: str) -> Dict[str, Any]:
    """Open a ZIP, list DSS files, classify them, return validation result.

    Two-level status reporting:
      ALERT_*  = genuine problem requiring manual review
      INFO_*   = multiple candidates but heuristic confidently selected one
      OK       = exactly one SV and one DV, no ambiguity
    """
    sc = scenario_id
    try:
        with zipfile.ZipFile(zip_path, "r") as zf:
            all_names = zf.namelist()
    except zipfile.BadZipFile:
        return {
            "dss_file_count": 0,
            "sv_candidate_count": 0,
            "dv_candidate_count": 0,
            "sv_selected": "",
            "dv_selected": "",
            "sv_reason": "",
            "dv_reason": "",
            "sv_all_candidates": "",
            "dv_all_candidates": "",
            "skipped_dss": "",
            "classification_method": "",
            "validation_status": "ALERT_BAD_ZIP",
        }

    dss_paths = [n for n in all_names if n.lower().endswith(".dss")]
    result = classify_dss_in_zip(dss_paths, scenario_id)

    sv_count = len(result["sv_candidates"])
    dv_count = len(result["dv_candidates"])

    # --- Detailed console logging ---
    log.info("[%s] Found %d DSS file(s) in ZIP (classified via %s):",
             sc, len(dss_paths), result["classification_method"])
    if result["skipped"]:
        for p in result["skipped"]:
            log.info("[%s]   SKIPPED (excluded subfolder): %s", sc, p)

    if sv_count > 0:
        log.info("[%s]   SV (input) candidates (%d):", sc, sv_count)
        for p in result["sv_candidates"]:
            marker = " <-- SELECTED" if p == result["sv_selected"] else ""
            log.info("[%s]     - %s%s", sc, os.path.basename(p), marker)
        log.info("[%s]   SV selection reason: %s", sc, result["sv_reason"])
    else:
        log.warning("[%s]   SV (input) candidates: NONE FOUND", sc)

    if dv_count > 0:
        log.info("[%s]   DV (output) candidates (%d):", sc, dv_count)
        for p in result["dv_candidates"]:
            marker = " <-- SELECTED" if p == result["dv_selected"] else ""
            log.info("[%s]     - %s%s", sc, os.path.basename(p), marker)
        log.info("[%s]   DV selection reason: %s", sc, result["dv_reason"])
    else:
        log.warning("[%s]   DV (output) candidates: NONE FOUND", sc)

    # --- Two-level status ---
    statuses = []
    if sv_count == 0:
        statuses.append("ALERT_NO_SV")
    elif sv_count > 1 and not result["sv_selected"]:
        statuses.append("ALERT_MULTIPLE_SV_NO_MATCH")
    elif sv_count > 1:
        statuses.append("INFO_MULTIPLE_SV")

    if dv_count == 0:
        statuses.append("ALERT_NO_DV")
    elif dv_count > 1 and not result["dv_selected"]:
        statuses.append("ALERT_MULTIPLE_DV_NO_MATCH")
    elif dv_count > 1:
        statuses.append("INFO_MULTIPLE_DV")

    if not statuses:
        statuses.append("OK")

    return {
        "dss_file_count": len(dss_paths),
        "sv_candidate_count": sv_count,
        "dv_candidate_count": dv_count,
        "sv_selected": result["sv_selected"] or "",
        "dv_selected": result["dv_selected"] or "",
        "sv_reason": result["sv_reason"],
        "dv_reason": result["dv_reason"],
        "sv_all_candidates": ";".join(result["sv_candidates"]),
        "dv_all_candidates": ";".join(result["dv_candidates"]),
        "skipped_dss": ";".join(result["skipped"]),
        "classification_method": result["classification_method"],
        "validation_status": "|".join(statuses),
    }


# ---------------------------------------------------------------------------
# Per-scenario worker
# ---------------------------------------------------------------------------
_print_lock = threading.Lock()


def process_scenario(
    scenario: Dict, s3_client, s3_bucket: str, dry_run: bool,
    rclone_remote: str,
) -> Dict[str, Any]:
    """Download, validate, and stage one scenario. Returns audit row dict."""
    sc = scenario["short_code"]
    folder_id = scenario["drive_folder_id"]
    row: Dict[str, Any] = {
        "scenario_id": sc,
        "study_name": scenario.get("study_name", ""),
        "run_date": scenario.get("run_date", ""),
        "drive_folder_id": folder_id,
        "zip_count": 0,
        "zip_selected": "",
        "zip_all_files": "",
        "zip_size_mb": "",
        "zip_drive_modified": "",
        "dss_file_count": 0,
        "classification_method": "",
        "sv_candidate_count": 0,
        "sv_selected": "",
        "sv_reason": "",
        "sv_all_candidates": "",
        "dv_candidate_count": 0,
        "dv_selected": "",
        "dv_reason": "",
        "dv_all_candidates": "",
        "skipped_dss": "",
        "trend_csv_count": 0,
        "trend_csv_selected": "",
        "trend_csv_all_files": "",
        "s3_staging_zip_key": "",
        "s3_staging_csv_key": "",
        "validation_status": "",
        "notes": scenario.get("notes", ""),
    }

    if not folder_id:
        row["validation_status"] = "NO_DRIVE_LINK"
        log.error("[%s] No Google Drive folder link found", sc)
        return row

    # --- List Model_Files and find ZIPs ---
    log.info("[%s] Listing Drive folder (ID: %s) ...", sc, folder_id[:12])
    all_model_files = rclone_lsjson(folder_id, "Model_Files/")
    zips = [f for f in all_model_files if f["Name"].lower().endswith(".zip")]
    row["zip_count"] = len(zips)
    row["zip_all_files"] = ";".join(sorted(f["Name"] for f in zips))

    if not zips:
        row["validation_status"] = "MISSING_ZIP"
        log.error("[%s] No ZIP files in Model_Files/", sc)
        return row

    pinned_zip = scenario.get("pinned_zip", "")
    if pinned_zip:
        match = [f for f in zips if f["Name"] == pinned_zip]
        if match:
            selected_zip = match[0]
            if len(zips) > 1:
                log.info("[%s] Pinned ZIP '%s' selected (%d total on Drive)",
                         sc, pinned_zip, len(zips))
        else:
            row["validation_status"] = "PINNED_ZIP_NOT_FOUND"
            log.error("[%s] Pinned ZIP '%s' not found among: %s",
                      sc, pinned_zip, ", ".join(f["Name"] for f in zips))
            return row
    else:
        if len(zips) > 1:
            log.warning("[%s] Multiple ZIPs found (%d): %s", sc, len(zips),
                        ", ".join(f["Name"] for f in zips))
        zips.sort(key=lambda f: f.get("ModTime", ""), reverse=True)
        selected_zip = zips[0]
    row["zip_selected"] = selected_zip["Name"]
    row["zip_drive_modified"] = selected_zip.get("ModTime", "")
    size_bytes = int(selected_zip.get("Size", 0))
    row["zip_size_mb"] = f"{size_bytes / (1024 * 1024):.1f}"

    # --- List Data_Extraction and find trend report CSVs ---
    trend_csvs: List[Dict] = []
    trend_files = rclone_lsjson(
        folder_id, "Data_Extraction/Variables_From_trend_report_variables_v5/"
    )
    trend_csvs = [
        f for f in trend_files
        if f["Name"].lower().endswith(".csv")
        and f["Name"].lower().startswith(sc.lower())
    ]

    row["trend_csv_count"] = len(trend_csvs)
    row["trend_csv_all_files"] = ";".join(sorted(f["Name"] for f in trend_csvs))

    selected_csv = None
    pinned_trend = scenario.get("pinned_trend", "")
    if not trend_csvs:
        log.warning("[%s] No trend report CSV found", sc)
    elif pinned_trend:
        match = [f for f in trend_csvs if f["Name"] == pinned_trend]
        if match:
            selected_csv = match[0]
            row["trend_csv_selected"] = selected_csv["Name"]
            if len(trend_csvs) > 1:
                log.info("[%s] Pinned trend CSV '%s' selected (%d total on Drive)",
                         sc, pinned_trend, len(trend_csvs))
        else:
            log.warning("[%s] Pinned trend CSV '%s' not found among: %s",
                        sc, pinned_trend, ", ".join(f["Name"] for f in trend_csvs))
    else:
        trend_csvs.sort(key=lambda f: f.get("ModTime", ""), reverse=True)
        selected_csv = trend_csvs[0]
        row["trend_csv_selected"] = selected_csv["Name"]
        if len(trend_csvs) > 1:
            log.warning("[%s] Multiple trend CSVs found (%d): %s", sc,
                        len(trend_csvs),
                        ", ".join(f["Name"] for f in trend_csvs))

    if dry_run:
        log.info("[%s] DRY RUN -- would download %s (%.1f MB) + %s",
                 sc, selected_zip["Name"], size_bytes / (1024 * 1024),
                 selected_csv["Name"] if selected_csv else "(no CSV)")
        row["validation_status"] = "DRY_RUN"
        return row

    # --- Download ZIP, validate, upload to S3 staging ---
    tmp_dir = tempfile.mkdtemp(prefix=f"gdrive_{sc}_")
    zip_filename = selected_zip["Name"]
    zip_local = os.path.join(tmp_dir, zip_filename)
    try:
        log.info("[%s] Downloading ZIP: %s (%.1f MB) ...", sc,
                 zip_filename, size_bytes / (1024 * 1024))
        ok = rclone_copy_file(folder_id, f"Model_Files/{zip_filename}", tmp_dir)
        if not ok or not os.path.exists(zip_local):
            row["validation_status"] = "DOWNLOAD_FAILED"
            log.error("[%s] ZIP download failed", sc)
            return row

        log.info("[%s] Validating ZIP ...", sc)
        val = validate_zip(zip_local, sc)
        row.update(val)

        if len(zips) > 1:
            existing_status = row["validation_status"]
            if "ALERT_MULTIPLE_ZIP" not in existing_status:
                row["validation_status"] = "ALERT_MULTIPLE_ZIP|" + existing_status
        if not selected_csv:
            existing_status = row["validation_status"]
            row["validation_status"] = existing_status + "|MISSING_TREND_REPORT"

        s3_zip_key = f"staging/{sc}/{zip_filename}"
        log.info("[%s] Uploading ZIP to s3://%s/%s ...", sc, s3_bucket, s3_zip_key)
        s3_client.upload_file(zip_local, s3_bucket, s3_zip_key)
        row["s3_staging_zip_key"] = s3_zip_key

    finally:
        if os.path.exists(zip_local):
            os.remove(zip_local)
        try:
            os.rmdir(tmp_dir)
        except OSError:
            pass

    # --- Download trend CSV and upload ---
    if selected_csv:
        csv_name = selected_csv["Name"]
        log.info("[%s] Downloading trend CSV: %s ...", sc, csv_name)
        csv_bytes = rclone_cat(
            folder_id,
            f"Data_Extraction/Variables_From_trend_report_variables_v5/{csv_name}",
        )
        if csv_bytes:
            s3_csv_key = f"staging/{sc}/{csv_name}"
            s3_client.put_object(Bucket=s3_bucket, Key=s3_csv_key, Body=csv_bytes)
            row["s3_staging_csv_key"] = s3_csv_key
            log.info("[%s] Uploaded trend CSV to s3://%s/%s",
                     sc, s3_bucket, s3_csv_key)
        else:
            log.warning("[%s] Failed to download trend CSV", sc)

    with _print_lock:
        status = row["validation_status"]
        if "ALERT" in status:
            log.warning("[%s] Done -- %s (review needed)", sc, status)
        elif "INFO" in status:
            log.info("[%s] Done -- %s (multiple candidates, selection confident)",
                     sc, status)
        else:
            log.info("[%s] Done -- %s", sc, status)

    return row


# ---------------------------------------------------------------------------
# Audit report
# ---------------------------------------------------------------------------
AUDIT_COLUMNS = [
    "scenario_id", "study_name", "run_date", "drive_folder_id",
    "zip_count", "zip_selected", "zip_all_files", "zip_size_mb",
    "zip_drive_modified", "dss_file_count", "classification_method",
    "sv_candidate_count", "sv_selected", "sv_reason", "sv_all_candidates",
    "dv_candidate_count", "dv_selected", "dv_reason", "dv_all_candidates",
    "skipped_dss",
    "trend_csv_count", "trend_csv_selected", "trend_csv_all_files",
    "s3_staging_zip_key", "s3_staging_csv_key",
    "validation_status", "notes",
]


def write_audit_report(rows: List[Dict], local_path: str,
                       s3_client, s3_bucket: str):
    """Write audit report to local CSV and upload to S3."""
    rows.sort(key=lambda r: r.get("scenario_id", ""))
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=AUDIT_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for r in rows:
        writer.writerow(r)
    csv_text = buf.getvalue()

    with open(local_path, "w") as f:
        f.write(csv_text)
    log.info("Audit report written to %s", local_path)

    s3_key = "staging/audit_report.csv"
    s3_client.put_object(Bucket=s3_bucket, Key=s3_key,
                         Body=csv_text.encode("utf-8"))
    log.info("Audit report uploaded to s3://%s/%s", s3_bucket, s3_key)

    print("\n" + "=" * 100)
    print("  DOWNLOAD & VALIDATION SUMMARY")
    print("=" * 100)
    ok = sum(1 for r in rows if r.get("validation_status") == "OK")
    info = sum(1 for r in rows
               if "INFO" in r.get("validation_status", "")
               and "ALERT" not in r.get("validation_status", ""))
    alerts = sum(1 for r in rows
                 if "ALERT" in r.get("validation_status", ""))
    missing = sum(1 for r in rows
                  if "MISSING" in r.get("validation_status", ""))
    dry = sum(1 for r in rows if r.get("validation_status") == "DRY_RUN")
    print(f"  Total scenarios:  {len(rows)}")
    print(f"  OK (clean):       {ok}")
    print(f"  INFO (multi, ok): {info}")
    print(f"  ALERT (review!):  {alerts}")
    print(f"  Missing files:    {missing}")
    if dry:
        print(f"  Dry run:          {dry}")

    # --- Detailed table ---
    print()
    hdr = "  {:<8} {:<3} {:<3} {:<30} {:<30} {}"
    row_fmt = "  {:<8} {:<3} {:<3} {:<30} {:<30} {}"
    print(hdr.format("Scenario", "SV#", "DV#", "SV selected", "DV selected", "Status"))
    print("  " + "-" * 96)
    for r in rows:
        sv_name = os.path.basename(r.get("sv_selected", ""))[:30] if r.get("sv_selected") else "-"
        dv_name = os.path.basename(r.get("dv_selected", ""))[:30] if r.get("dv_selected") else "-"
        print(row_fmt.format(
            r.get("scenario_id", ""),
            str(r.get("sv_candidate_count", "")),
            str(r.get("dv_candidate_count", "")),
            sv_name,
            dv_name,
            r.get("validation_status", ""),
        ))

    # --- Print scenarios needing attention ---
    attention = [r for r in rows if "ALERT" in r.get("validation_status", "")]
    if attention:
        print()
        print("  SCENARIOS REQUIRING MANUAL REVIEW:")
        print("  " + "-" * 96)
        for r in attention:
            print(f"  {r.get('scenario_id', '')}:")
            print(f"    Status:  {r.get('validation_status', '')}")
            if r.get("sv_all_candidates"):
                print(f"    SV candidates: {r['sv_all_candidates']}")
            if r.get("dv_all_candidates"):
                print(f"    DV candidates: {r['dv_all_candidates']}")
            if r.get("skipped_dss"):
                print(f"    Skipped DSS:   {r['skipped_dss']}")

    print("=" * 100 + "\n")


# ---------------------------------------------------------------------------
# Commands
# ---------------------------------------------------------------------------
def cmd_download(args):
    """Download, validate, and stage scenarios to S3."""
    global RCLONE_REMOTE
    RCLONE_REMOTE = args.rclone_remote

    fmt = _detect_csv_format(args.listing_csv)
    if fmt == "source":
        scenarios = read_scenario_source_csv(args.listing_csv)
        log.info("Using model_run_file_source.csv format")
        if not args.include_all:
            before = len(scenarios)
            scenarios = [s for s in scenarios if s["download_status"] == "ready"]
            skipped = before - len(scenarios)
            if skipped:
                log.info("Filtered to 'ready' scenarios (%d skipped; use --include-all to download all)", skipped)
    elif fmt == "excel":
        scenarios = read_scenario_listing(args.listing_csv, args.sheet)
        log.info("Using legacy Excel listing format")
    else:
        scenarios = read_scenario_listing_csv(args.listing_csv)
        log.info("Using legacy v6 CSV listing format")

    if args.scenarios:
        filter_set = set(s.lower() for s in args.scenarios)
        scenarios = [s for s in scenarios if s["short_code"].lower() in filter_set]

    if not scenarios:
        log.error("No scenarios matched the filter")
        return

    log.info("Found %d scenarios to process", len(scenarios))
    for s in scenarios:
        fid = s["drive_folder_id"]
        log.info("  %s: %s (folder: %s)", s["short_code"],
                 s.get("study_name") or s["drive_folder_name"],
                 fid[:12] + "..." if fid else "NONE")

    s3_client = boto3.client("s3")
    results: List[Dict] = []
    workers = args.workers

    if workers <= 1:
        for sc in scenarios:
            row = process_scenario(sc, s3_client, args.s3_bucket,
                                   args.dry_run, RCLONE_REMOTE)
            results.append(row)
    else:
        with ThreadPoolExecutor(max_workers=workers) as pool:
            futures = {
                pool.submit(process_scenario, sc, s3_client,
                            args.s3_bucket, args.dry_run, RCLONE_REMOTE): sc
                for sc in scenarios
            }
            for future in as_completed(futures):
                sc = futures[future]
                try:
                    row = future.result()
                    results.append(row)
                except Exception:
                    log.exception("[%s] Worker failed", sc["short_code"])
                    results.append({
                        "scenario_id": sc["short_code"],
                        "study_name": sc.get("study_name", ""),
                        "validation_status": "WORKER_ERROR",
                        "notes": sc.get("notes", ""),
                    })

    output_dir = Path(getattr(args, "output_dir", None) or DEFAULT_OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)
    local_report = str(output_dir / "audit_report.csv")
    write_audit_report(results, local_report, s3_client, args.s3_bucket)


def cmd_promote(args):
    """Copy files from staging/<shortcode>/ to ready/<shortcode>/."""
    s3 = boto3.client("s3")
    bucket = args.s3_bucket

    staging_objects: List[Dict] = []
    paginator = s3.get_paginator("list_objects_v2")
    for page in paginator.paginate(Bucket=bucket, Prefix="staging/"):
        for obj in page.get("Contents", []):
            staging_objects.append(obj)

    groups: Dict[str, List[str]] = {}
    for obj in staging_objects:
        key = obj["Key"]
        parts = key.split("/")
        if len(parts) >= 3 and parts[0] == "staging":
            sc = parts[1]
            if re.match(r"^s\d{4}$", sc):
                groups.setdefault(sc, []).append(key)

    filter_ids = None
    if args.scenarios:
        filter_ids = {s.strip().lower() for s in args.scenarios.split(",")}

    if filter_ids:
        groups = {k: v for k, v in groups.items() if k in filter_ids}
        if not groups:
            log.error("No staged files found for %s",
                      ", ".join(sorted(filter_ids)))
            return

    if not groups:
        log.error("No staged scenarios found in s3://%s/staging/", bucket)
        return

    print(f"\nAbout to promote {len(groups)} scenario(s) from staging/ to ready/:")
    for sc in sorted(groups):
        files = [k.split("/")[-1] for k in groups[sc]]
        print(f"  {sc}: {', '.join(files)}")

    confirm = input("\nContinue? [y/N] ").strip().lower()
    if confirm != "y":
        print("Aborted.")
        return

    for sc in sorted(groups):
        for src_key in groups[sc]:
            filename = src_key.split("/")[-1]
            dest_key = f"ready/{sc}/{filename}"
            log.info("Copying s3://%s/%s -> s3://%s/%s",
                     bucket, src_key, bucket, dest_key)
            s3.copy_object(
                Bucket=bucket,
                CopySource={"Bucket": bucket, "Key": src_key},
                Key=dest_key,
            )
        log.info("[%s] Promoted to ready/", sc)

    print(f"\nDone. Promoted {len(groups)} scenario(s) to ready/.")
    print("The Lambda will trigger on each ZIP upload.")


# ---------------------------------------------------------------------------
# Scan subcommand — CSV-based path validation and Drive content auditing
# ---------------------------------------------------------------------------

SCAN_AUDIT_COLUMNS = [
    "scenario_id", "hydroclimate", "drive_folder_name", "drive_folder_id",
    "folder_name_match", "access_mode", "model_files_link",
    "zip_count", "zip_names", "zip_selected", "zip_size_mb",
    "trend_csv_count", "trend_csv_names", "trend_csv_selected",
    "expected_dv_filename", "expected_sv_filename",
    "dv_root", "status",
]


def read_scenario_source_csv(csv_path: str) -> List[Dict[str, str]]:
    """Read model_run_file_source.csv (our curated source of truth).

    Columns: short_code, drive_folder_id, drive_folder_name,
             pinned_model_run_zip, pinned_trend_csv, download_status, notes
    """
    scenarios = []
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            short_code = row.get("short_code", "").strip()
            if not short_code:
                continue
            folder_name = row.get("drive_folder_name", "").strip()
            scenarios.append({
                "short_code": short_code,
                "drive_folder_id": row.get("drive_folder_id", "").strip(),
                "drive_folder_name": folder_name,
                "pinned_zip": row.get("pinned_model_run_zip", "").strip(),
                "pinned_trend": row.get("pinned_trend_csv", "").strip(),
                "download_status": row.get("download_status", "").strip(),
                "notes": row.get("notes", "").strip(),
                "dv_root": folder_name,
                "model_files_link": "",
                "hydroclimate": "",
                "dv_filename": "",
                "sv_filename": "",
                "study_name": "",
                "run_date": "",
            })
    return scenarios


def read_scenario_listing_csv(csv_path: str) -> List[Dict[str, str]]:
    """Read external team's v6 scenario listing CSV (legacy format).

    Kept for backward compatibility. Prefer read_scenario_source_csv().
    """
    scenarios = []
    with open(csv_path, encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        for row in reader:
            short_code = row.get("Index", "").strip()
            if not short_code:
                continue

            model_link = row.get("ModelFilesLink", "").strip()
            folder_id = ""
            m = RE_FOLDER_ID.search(model_link)
            if m:
                folder_id = m.group(1)

            dv_path = row.get("DV_Path", "").strip().replace("\\", "/")
            sv_path = row.get("SV_Path", "").strip().replace("\\", "/")

            dv_filename = dv_path.rsplit("/", 1)[-1] if dv_path else ""
            sv_filename = sv_path.rsplit("/", 1)[-1] if sv_path else ""
            dv_root = dv_path.split("/Model_Files/")[0] if "/Model_Files/" in dv_path else ""

            scenarios.append({
                "short_code": short_code,
                "study_name": row.get("StudyName", "").strip(),
                "drive_folder_name": row.get("GoogleDriveFolderName", "").strip(),
                "drive_folder_id": folder_id,
                "model_files_link": model_link,
                "hydroclimate": row.get("HydroClimate", "").strip(),
                "dv_filename": dv_filename,
                "sv_filename": sv_filename,
                "dv_root": dv_root,
            })
    return scenarios


def _detect_csv_format(path: str) -> str:
    """Return 'source' for model_run_file_source.csv, 'listing' for legacy CSV,
    or 'excel' for .xlsx files."""
    if path.lower().endswith(".xlsx"):
        return "excel"
    with open(path, encoding="utf-8-sig") as f:
        header = f.readline()
    return "source" if "short_code" in header else "listing"


def _resolve_drive_access(scenario: Dict) -> tuple:
    """Determine how to access a scenario's Drive folder.

    Returns (folder_id_or_empty, model_files_path, trend_path, access_mode).
    access_mode is "id" (folder ID) or "path" (Shared Drive path fallback).
    """
    folder_id = scenario["drive_folder_id"]
    folder_name = scenario["drive_folder_name"]
    dv_root = scenario["dv_root"]

    if folder_id:
        return (
            folder_id,
            "Model_Files/",
            "Data_Extraction/Variables_From_trend_report_variables_v5/",
            "id",
        )

    # No folder ID -- fall back to Shared Drive path navigation.
    # Prefer dv_root (from DV_Path) since it's the actual folder name on Drive;
    # fall back to GoogleDriveFolderName if dv_root is empty.
    base = dv_root or folder_name
    if not base:
        return ("", "", "", "none")

    return (
        "",
        f"{base}/Model_Files/",
        f"{base}/Data_Extraction/Variables_From_trend_report_variables_v5/",
        "path",
    )


def scan_scenario(
    scenario: Dict, rclone_remote: str,
) -> Dict[str, Any]:
    """List Drive contents for one scenario, report zip/csv counts.

    Navigates by folder ID when available; falls back to Shared Drive
    path navigation for scenarios whose ModelFilesLink has no URL.
    When folder_name != dv_root, tries both paths and reports which worked.
    """
    sc = scenario["short_code"]
    folder_id = scenario["drive_folder_id"]
    folder_name = scenario["drive_folder_name"]
    dv_root = scenario["dv_root"]

    row: Dict[str, Any] = {
        "scenario_id": sc,
        "hydroclimate": scenario["hydroclimate"],
        "drive_folder_name": folder_name,
        "drive_folder_id": folder_id[:16] + "..." if len(folder_id) > 16 else folder_id,
        "folder_name_match": "",
        "model_files_link": scenario["model_files_link"][:60],
        "access_mode": "",
        "zip_count": 0,
        "zip_names": "",
        "zip_selected": "",
        "zip_size_mb": "",
        "trend_csv_count": 0,
        "trend_csv_names": "",
        "trend_csv_selected": "",
        "expected_dv_filename": scenario["dv_filename"],
        "expected_sv_filename": scenario["sv_filename"],
        "dv_root": dv_root,
        "status": "",
    }

    if dv_root and dv_root != folder_name:
        row["folder_name_match"] = "MISMATCH"
        log.info("[%s] Folder name mismatch: GoogleDriveFolderName='%s' vs DV_Path root='%s'",
                 sc, folder_name, dv_root)
    elif dv_root:
        row["folder_name_match"] = "OK"
    else:
        row["folder_name_match"] = "NO_DV_PATH"

    fid, model_path, trend_path, access_mode = _resolve_drive_access(scenario)
    row["access_mode"] = access_mode

    if access_mode == "none":
        row["status"] = "NO_DRIVE_ACCESS"
        log.warning("[%s] No folder ID and no folder name -- cannot scan", sc)
        return row

    if access_mode == "path":
        log.info("[%s] No folder ID; using Shared Drive path: %s", sc, model_path)
    else:
        log.info("[%s] Using folder ID: %s ...", sc, folder_id[:12])

    statuses = []

    # --- List Model_Files/ for ZIPs ---
    model_files = rclone_lsjson(fid, model_path, rclone_remote=rclone_remote)
    zips = [f for f in model_files if f["Name"].lower().endswith(".zip")]
    row["zip_count"] = len(zips)
    row["zip_names"] = ";".join(sorted(f["Name"] for f in zips))

    pinned_zip = scenario.get("pinned_zip", "")
    if not zips:
        statuses.append("MISSING_ZIP")
        log.warning("[%s] No ZIP files found in Model_Files/", sc)
    elif pinned_zip:
        match = [f for f in zips if f["Name"] == pinned_zip]
        if match:
            row["zip_selected"] = match[0]["Name"]
            row["zip_size_mb"] = f"{int(match[0].get('Size', 0)) / (1024 * 1024):.1f}"
            if len(zips) > 1:
                log.info("[%s] Pinned ZIP '%s' selected (%d total on Drive)",
                         sc, pinned_zip, len(zips))
        else:
            statuses.append("PINNED_ZIP_NOT_FOUND")
            log.warning("[%s] Pinned ZIP '%s' not found among: %s",
                        sc, pinned_zip, ", ".join(f["Name"] for f in zips))
    else:
        zips.sort(key=lambda f: f.get("ModTime", ""), reverse=True)
        selected = zips[0]
        row["zip_selected"] = selected["Name"]
        row["zip_size_mb"] = f"{int(selected.get('Size', 0)) / (1024 * 1024):.1f}"
        if len(zips) > 1:
            statuses.append("ALERT_MULTIPLE_ZIP")
            log.warning("[%s] Multiple ZIPs (%d): %s", sc, len(zips),
                        ", ".join(f["Name"] for f in zips))

    # --- List Data_Extraction/ for trend report CSVs ---
    trend_files = rclone_lsjson(fid, trend_path, rclone_remote=rclone_remote)
    trend_csvs = [
        f for f in trend_files
        if f["Name"].lower().endswith(".csv")
        and f["Name"].lower().startswith(sc.lower())
    ]
    row["trend_csv_count"] = len(trend_csvs)
    row["trend_csv_names"] = ";".join(sorted(f["Name"] for f in trend_csvs))

    pinned_trend = scenario.get("pinned_trend", "")
    if not trend_csvs:
        statuses.append("MISSING_TREND")
        log.warning("[%s] No trend report CSV found", sc)
    elif pinned_trend:
        match = [f for f in trend_csvs if f["Name"] == pinned_trend]
        if match:
            row["trend_csv_selected"] = match[0]["Name"]
            if len(trend_csvs) > 1:
                log.info("[%s] Pinned trend CSV '%s' selected (%d total on Drive)",
                         sc, pinned_trend, len(trend_csvs))
        else:
            statuses.append("PINNED_TREND_NOT_FOUND")
            log.warning("[%s] Pinned trend CSV '%s' not found among: %s",
                        sc, pinned_trend, ", ".join(f["Name"] for f in trend_csvs))
    else:
        trend_csvs.sort(key=lambda f: f.get("ModTime", ""), reverse=True)
        row["trend_csv_selected"] = trend_csvs[0]["Name"]
        if len(trend_csvs) > 1:
            statuses.append("ALERT_MULTIPLE_TREND")
            log.warning("[%s] Multiple trend CSVs (%d): %s", sc, len(trend_csvs),
                        ", ".join(f["Name"] for f in trend_csvs))

    if dv_root and dv_root != folder_name:
        statuses.append("FOLDER_MISMATCH")

    row["status"] = "|".join(statuses) if statuses else "OK"

    log.info("[%s] Scan done -- %s (access=%s, zips=%d, trend_csvs=%d)",
             sc, row["status"], access_mode, row["zip_count"], row["trend_csv_count"])
    return row


def write_scan_audit(rows: List[Dict], local_path: str):
    """Write scan audit CSV to disk."""
    rows.sort(key=lambda r: r.get("scenario_id", ""))
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=SCAN_AUDIT_COLUMNS,
                            extrasaction="ignore")
    writer.writeheader()
    for r in rows:
        writer.writerow(r)
    csv_text = buf.getvalue()

    with open(local_path, "w") as f:
        f.write(csv_text)
    log.info("Scan audit written to %s", local_path)

    print("\n" + "=" * 100)
    print("  SCAN AUDIT SUMMARY")
    print("=" * 100)
    ok = sum(1 for r in rows if r.get("status") == "OK")
    no_id = sum(1 for r in rows if r.get("status") == "NO_FOLDER_ID")
    alerts = sum(1 for r in rows if "ALERT" in r.get("status", ""))
    missing = sum(1 for r in rows if "MISSING" in r.get("status", ""))
    mismatch = sum(1 for r in rows if "FOLDER_MISMATCH" in r.get("status", ""))
    print(f"  Total scenarios:       {len(rows)}")
    print(f"  OK (clean):            {ok}")
    print(f"  No folder ID:          {no_id}")
    print(f"  Alerts (multi files):  {alerts}")
    print(f"  Missing files:         {missing}")
    print(f"  Folder mismatches:     {mismatch}")

    print()
    hdr = "  {:<8} {:<26} {:<5} {:>4} {:>4} {:<8} {}"
    print(hdr.format("Scenario", "Hydroclimate", "Via", "Zips", "CSVs", "Match", "Status"))
    print("  " + "-" * 96)
    for r in rows:
        hydro = r.get("hydroclimate", "")[:26]
        print(hdr.format(
            r.get("scenario_id", ""),
            hydro,
            r.get("access_mode", "")[:5],
            str(r.get("zip_count", "")),
            str(r.get("trend_csv_count", "")),
            r.get("folder_name_match", ""),
            r.get("status", ""),
        ))

    attention = [r for r in rows
                 if r.get("status", "") not in ("OK", "NO_FOLDER_ID", "LOCAL_ONLY")]
    if attention:
        print()
        print("  SCENARIOS REQUIRING ATTENTION:")
        print("  " + "-" * 90)
        for r in attention:
            print(f"  {r['scenario_id']}: {r['status']}")
            if r.get("zip_names"):
                print(f"    ZIPs: {r['zip_names']}")
            if r.get("trend_csv_names"):
                print(f"    Trend CSVs: {r['trend_csv_names']}")
            if r.get("folder_name_match") == "MISMATCH":
                print(f"    Folder name: {r['drive_folder_name']}")
                print(f"    DV_Path root: {r['dv_root']}")

    print("=" * 100 + "\n")


def cmd_scan(args):
    """Scan Drive contents using scenario_source.csv or legacy v6 CSV."""
    global RCLONE_REMOTE
    RCLONE_REMOTE = args.rclone_remote

    fmt = _detect_csv_format(args.listing_csv)
    if fmt == "source":
        scenarios = read_scenario_source_csv(args.listing_csv)
        log.info("Using scenario_source.csv format")
        if not args.include_all:
            before = len(scenarios)
            scenarios = [s for s in scenarios if s["download_status"] == "ready"]
            skipped = before - len(scenarios)
            if skipped:
                log.info("Filtered to 'ready' scenarios (%d skipped; use --include-all to scan all)", skipped)
    else:
        scenarios = read_scenario_listing_csv(args.listing_csv)
        log.info("Using legacy v6 listing CSV format")

    if args.scenarios:
        filter_set = set(s.lower() for s in args.scenarios)
        scenarios = [s for s in scenarios if s["short_code"].lower() in filter_set]

    if not scenarios:
        log.error("No scenarios found in listing")
        return

    log.info("Loaded %d scenarios from CSV", len(scenarios))

    if args.local_only:
        log.info("Local-only mode: writing manifest without Drive access")
        results = []
        for sc in scenarios:
            row = {
                "scenario_id": sc["short_code"],
                "hydroclimate": sc["hydroclimate"],
                "drive_folder_name": sc["drive_folder_name"],
                "drive_folder_id": sc["drive_folder_id"][:16] + "..."
                    if len(sc["drive_folder_id"]) > 16 else sc["drive_folder_id"],
                "folder_name_match": "MISMATCH" if sc["dv_root"] and sc["dv_root"] != sc["drive_folder_name"]
                    else ("OK" if sc["dv_root"] else "NO_DV_PATH"),
                "access_mode": "id" if sc["drive_folder_id"] else (
                    "path" if (sc["dv_root"] or sc["drive_folder_name"]) else "none"),
                "model_files_link": sc["model_files_link"][:60],
                "zip_count": "",
                "zip_names": "",
                "zip_selected": "",
                "zip_size_mb": "",
                "trend_csv_count": "",
                "trend_csv_names": "",
                "trend_csv_selected": "",
                "expected_dv_filename": sc["dv_filename"],
                "expected_sv_filename": sc["sv_filename"],
                "dv_root": sc["dv_root"],
                "status": "LOCAL_ONLY" + ("|FOLDER_MISMATCH"
                    if sc["dv_root"] and sc["dv_root"] != sc["drive_folder_name"] else "")
                    + ("|NO_FOLDER_ID" if not sc["drive_folder_id"] else ""),
            }
            results.append(row)
        local_path = os.path.join(os.getcwd(), "scan_manifest.csv")
        write_scan_audit(results, local_path)
        return

    results: List[Dict] = []
    workers = args.workers

    if workers <= 1:
        for sc in scenarios:
            row = scan_scenario(sc, RCLONE_REMOTE)
            results.append(row)
    else:
        with ThreadPoolExecutor(max_workers=workers) as pool:
            futures = {
                pool.submit(scan_scenario, sc, RCLONE_REMOTE): sc
                for sc in scenarios
            }
            for future in as_completed(futures):
                sc = futures[future]
                try:
                    row = future.result()
                    results.append(row)
                except Exception:
                    log.exception("[%s] Scan worker failed", sc["short_code"])
                    results.append({
                        "scenario_id": sc["short_code"],
                        "status": "WORKER_ERROR",
                    })

    output_dir = Path(getattr(args, "output_dir", None) or DEFAULT_OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)
    local_path = str(output_dir / "scan_audit.csv")
    write_scan_audit(results, local_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)-5s %(message)s",
        datefmt="%Y-%m-%dT%H:%M:%SZ",
    )

    parser = argparse.ArgumentParser(
        description="Bulk download model runs from Google Drive to S3 via rclone"
    )
    sub = parser.add_subparsers(dest="command", required=True)

    dl = sub.add_parser("download", help="Download, validate, and stage to S3")
    dl.add_argument("--listing-csv", required=True,
                    help="Path to model_run_file_source.csv (or legacy Excel listing)")
    dl.add_argument("--sheet", default="HistHydro_20260223",
                    help="Excel sheet name (only for legacy .xlsx format)")
    dl.add_argument("--s3-bucket", required=True,
                    help="S3 bucket for staging (e.g., coeqwal-model-run)")
    dl.add_argument("--workers", type=int, default=4,
                    help="Number of concurrent download workers (default: 4)")
    dl.add_argument("--scenarios", nargs="*",
                    help="Optional: only process these scenario short codes")
    dl.add_argument("--rclone-remote", default="gdrive",
                    help="Name of the rclone remote (default: gdrive)")
    dl.add_argument("--include-all", action="store_true",
                    help="Include needs_review/skip scenarios (default: ready only)")
    dl.add_argument("--dry-run", action="store_true",
                    help="List files without downloading")
    dl.add_argument("--output-dir", default=None,
                    help=f"Directory for audit_report.csv (default: "
                         f"{DEFAULT_OUTPUT_DIR}). Auto-created if missing.")

    pr = sub.add_parser("promote",
                        help="Copy staged files from staging/ to ready/")
    pr.add_argument("--s3-bucket", required=True,
                    help="S3 bucket (e.g., coeqwal-model-run)")
    pr.add_argument("--scenarios",
                    help="Comma-separated scenario IDs to promote (default: all in staging)")

    sc = sub.add_parser("scan",
                        help="Scan Drive contents using scenario_source.csv")
    sc.add_argument("--listing-csv", required=True,
                    help="Path to scenario_source.csv (or legacy v6 CSV)")
    sc.add_argument("--rclone-remote", default="gdrive",
                    help="Name of the rclone remote (default: gdrive)")
    sc.add_argument("--workers", type=int, default=4,
                    help="Number of concurrent scan workers (default: 4)")
    sc.add_argument("--scenarios", nargs="*",
                    help="Optional: only scan these scenario short codes")
    sc.add_argument("--include-all", action="store_true",
                    help="Include needs_review/skip scenarios (default: ready only)")
    sc.add_argument("--local-only", action="store_true",
                    help="Parse CSV and write manifest without Drive access")
    sc.add_argument("--output-dir", default=None,
                    help=f"Directory for scan_audit.csv (default: "
                         f"{DEFAULT_OUTPUT_DIR}). Auto-created if missing.")

    args = parser.parse_args()
    if args.command == "download":
        cmd_download(args)
    elif args.command == "promote":
        cmd_promote(args)
    elif args.command == "scan":
        cmd_scan(args)


if __name__ == "__main__":
    main()
